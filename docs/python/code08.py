from openpyxl import Workbook
wb = Workbook()
wb.create_sheet()
wb.create_sheet()
wb.create_sheet()

for sheet in wb:
    print(sheet.title)
