import datetime
from openpyxl import Workbook
wb = Workbook(guess_types=True)
ws = wb.active
# set date using a Python datetime
ws['A1'] = datetime.datetime(2010, 7, 21)
 
print(ws['A1'].number_format)
 
# set percentage using a string followed by the percent sign
ws['B1'] = '3.14%'

print(ws['B1'].value)

print(ws['B1'].number_format)
