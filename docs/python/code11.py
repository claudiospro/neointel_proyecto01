from openpyxl import load_workbook

wb = load_workbook('tmp/sample_book.xltm', keep_vba=True) 
ws = wb.active 
ws['D2'] = 42 

wb.save('tmp/sample_book.xlsm') 

# or you can overwrite the current document template
# wb.save('tmp/sample_book.xltm')
