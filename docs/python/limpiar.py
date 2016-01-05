from openpyxl import load_workbook
from openpyxl import Workbook
import re

patron1 = re.compile('\+[0-9][0-9][0-9]')
patron2 = re.compile('\+[0-9][0-9]')

wb10  = load_workbook('tmp/limpiar-old.xlsx')
ws11 = wb10['huecos']
ws12 = wb10['alava']

wb20  = Workbook()
ws21 = wb20.active
ws21.title = 'huecos'
ws22 = wb20.create_sheet()
ws22.title = 'alava'

list_21 = list()
list_22 = list()

ws21['A1'] = 'municipio'
ws21['B1'] = 'direccion'
ws21['C1'] = 'numero'

i = 1 
j = 1

for row in ws11.rows:
    if i == 1:
        i+=1
        j+=1
        continue
    index_21 = ''

    municipio1 = ws11.cell(row = i, column = 4)
    index_21+=municipio1.value
    calle1 = ws11.cell(row = i, column = 6)
    index_21+=calle1.value
    numero1 = ws11.cell(row = i, column = 7)
    index_21+=numero1.value

    if ( (index_21 in list_21) == False ):    
        list_21.append(index_21)
        municipio2 = ws21.cell(row = j, column = 1)
        municipio2.value = municipio1.value
        calle2 = ws21.cell(row = j, column = 2)
        calle2.value = calle1.value
        numero2 = ws21.cell(row = j, column = 3)
        numero2.value = numero1.value
        j+=1        
    i+=1

ws22['A1'] = 'provincia'
ws22['B1'] = 'ciudad'
ws22['C1'] = 'direccion'
ws22['D1'] = 'numero'
ws22['E1'] = 'telefono'

i = 1 
j = 1
for row in ws12.rows:
    if i == 1:
        i+=1
        continue
    provincia1 = ws12.cell(row = i, column = 5)
    provincia2 = ws22.cell(row = i, column = 1)
    provincia2.value = provincia1.value

    ciudad1 = ws12.cell(row = i, column = 4)
    ciudad2 = ws22.cell(row = i, column = 2)
    ciudad2.value = ciudad1.value

    direccion1 = ws12.cell(row = i, column = 2)
    direccion2 = ws22.cell(row = i, column = 3)
    numero2 = ws22.cell(row = i, column = 4)
    l = [pos for pos, char in enumerate(direccion1.value) if char == ',']
    if len(l) > 0:
        direccion2.value = direccion1.value[0:l[-1]]
        numero2.value = int(direccion1.value[l[-1]+1:])
    else:
        direccion2.value = direccion1.value
        if i == 1:
            numero2.value = 'Numer3o'
        else:            
            numero2.value = ''
    
    telefono1 = ws12.cell(row = i, column = 6)
    telefono2 = ws22.cell(row = i, column = 5)
    telefono2.value = telefono1.value
    telefono2.value = patron1.sub('', telefono2.value)
    telefono2.value = patron2.sub('', telefono2.value)
    telefono2.value = telefono2.value.replace('.', '') 
    telefono2.value = telefono2.value.replace(' ', '')
    telefono2.value = int(telefono2.value)
    i+=1

wb20.save(filename = 'tmp/limpiar-new.xlsx')
