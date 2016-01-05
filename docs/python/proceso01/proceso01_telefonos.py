import sys
import re
from openpyxl import load_workbook
from openpyxl import Workbook
import config.cols 

cols = config.cols.telefonos

wb10  = load_workbook(sys.argv[1])
ws12 = wb10.active

wb20  = Workbook()
ws22 = wb20.active

ws22['A1'] = 'provincia'
ws22['B1'] = 'ciudad'
ws22['C1'] = 'direccion'
ws22['D1'] = 'numero'
ws22['E1'] = 'telefono'

i  = 1 

for row in ws12.rows:
    if i == 1:
        i+=1
        continue
    provincia1 = ws12.cell(row = i, column = cols['provincia'])
    ciudad1 = ws12.cell(row = i, column = cols['ciudad'])
    direccion1 = ws12.cell(row = i, column = cols['direccion'])
    telefono1 = ws12.cell(row = i, column = cols['telefono'])

    provincia2 = ws22.cell(row = i, column = 1)
    provincia2.value = provincia1.value
    
    ciudad2 = ws22.cell(row = i, column = 2)
    ciudad2.value = ciudad1.value

    direccion2 = ws22.cell(row = i, column = 3)
    numero2 = ws22.cell(row = i, column = 4)

    l = [pos for pos, char in enumerate(direccion1.value) if char == ',']
    if len(l) == 0:
        direccion2.value = direccion1.value
        if i == 1:
            numero2.value = 'Numero'
        else:            
            numero2.value = ''
    elif len(l) == 1:
        direccion2.value = direccion1.value[0:l[-1]]
        numero2.value = direccion1.value[l[-1]+1:]
    elif len(l) > 1:
        direccion2.value = direccion1.value
        print 'problema con registro: "' + str(i) + '" value: "' + str(direccion2.value) + '"'
    
    patron1 = re.compile("\+[0-9][0-9][0-9]")
    patron2 = re.compile("\+[0-9][0-9]")
    telefono2 = ws22.cell(row = i, column = 5)
    telefono2.value = telefono1.value
    telefono2.value = patron1.sub('', telefono2.value)
    telefono2.value = patron2.sub('', telefono2.value)
    telefono2.value = telefono2.value.replace('.', '') 
    telefono2.value = telefono2.value.replace(' ', '')
    
    i+=1

wb20.save(filename = 'procesado_'+sys.argv[1])
