import sys
import re
from openpyxl import load_workbook
from openpyxl import Workbook

import config.cols 
cols = config.cols.huecos

wb10 = load_workbook(sys.argv[1])
ws11 = wb10.active

wb20  = Workbook()
ws21 = wb20.active

ws21['A1'] = 'municipio'
ws21['B1'] = 'direccion'
ws21['C1'] = 'numero'

list_21 = list()
i = j = 1 

for row in ws11.rows:
    if i == 1:
        i+=1
        j+=1
        continue
    index_21 = ''

    municipio1 = ws11.cell(row = i, column = huecos['municipio'])
    index_21+=municipio1.value
    calle1 = ws11.cell(row = i, column = huecos['direccion'])
    index_21+=calle1.value
    numero1 = ws11.cell(row = i, column = huecos['numero'])
    index_21 += numero1.value

    if ( (index_21 in list_21) == False ):    
        list_21.append(index_21)
        municipio2 = ws21.cell(row = j, column = 1)
        municipio2.value = municipio1.value
        calle2 = ws21.cell(row = j, column = 2)
        calle2.value = calle1.value
        numero2 = ws21.cell(row = j, column = 3)
        numero2.value = numero1.value
        j+=1
    i+=1

wb20.save(filename = 'procesado_'+sys.argv[1])
