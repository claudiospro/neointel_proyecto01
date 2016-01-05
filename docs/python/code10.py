from openpyxl import load_workbook

wb = load_workbook('tmp/empty_book.xlsx') 
ws = wb.active 
ws['D2'] = 42 

wb.save('tmp/empty_book.xlsx') 
wb.save('tmp/empty_book.xltx')
