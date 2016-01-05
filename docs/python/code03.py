from openpyxl import Workbook
wb = Workbook()
ws1 = wb.create_sheet() # insert at the end (default)
ws2 = wb.create_sheet(0) # insert at first position
