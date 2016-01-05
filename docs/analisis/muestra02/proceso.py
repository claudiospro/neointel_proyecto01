import sys
import re
import time
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.cell import column_index_from_string

import config.cols 
cols = config.cols.huecos

wb10 = load_workbook(sys.argv[1])
ws11 = wb10['X']

wb20  = Workbook()
ws21 = wb20.active

ws21['A1'] = 'municipio'
ws21['B1'] = 'direccion'

list_21 = list()
i = j = 1 

# print ws11.max_row

for row in ws11.rows:
    if i == 1:
        i+=1
        j+=1
        continue
    index_21 = ''

    municipio1 = ws11.cell(row = i, column = cols['municipio'])
    direcion1 = ws11.cell(row = i, column = cols['direccion'])

    if (municipio1.value is not None and direcion1.value is not None):
        index_21+=municipio1.value
        index_21+=direcion1.value
        # print index_21

        if ( (index_21 in list_21) == False ):    
            list_21.append(index_21)
            municipio2 = ws21.cell(row = j, column = 1)
            direcion2 = ws21.cell(row = j, column = 2)
            
            municipio2.value = municipio1.value
            municipio2.value = municipio2.value.decode('utf-8') 
            print type(municipio2.value)
            direcion2.value = direcion1.value
            print type(direcion2.value)
            j+=1
    i+=1

wb20.save(filename = 'procesado_calle_'+sys.argv[1])
